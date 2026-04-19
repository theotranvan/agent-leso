"""Tests unitaires pour SafResultsParser (double-check analytique qL²/8)."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.connectors.structural.base import AnomalyLevel
from app.connectors.structural.results_parser import (
    DIVERGENCE_INFO_THRESHOLD_PCT,
    DIVERGENCE_WARNING_THRESHOLD_PCT,
    SafResultsParser,
)


# Modèle structurel de test : 1 poutre sur 2 appuis, L=6m, chargement réparti
MODEL_SIMPLE_BEAM = {
    "nodes": [
        {"id": "N1", "x": 0.0, "y": 0.0, "z": 3.0},
        {"id": "N2", "x": 6.0, "y": 0.0, "z": 3.0},
    ],
    "members": [
        {"id": "B1", "type": "beam", "node_start": "N1", "node_end": "N2",
         "section": "POU_30x50", "material": "C30/37"},
    ],
    "supports": [
        {"id": "S1", "node": "N1", "type": "pinned"},
        {"id": "S2", "node": "N2", "type": "roller"},
    ],
    "loads": [
        {"id": "L1", "case": "Q1", "target": "B1", "target_type": "member",
         "type": "uniform_vertical", "direction": "-Z",
         "value_kN_m": 10.0, "category": "Variable"},
    ],
}


def _write_csv_results(tmp_path: Path, member_id: str, m_kNm: float, utilization: float = 0.8) -> Path:
    """Produit un CSV résultats pour un seul élément."""
    import csv
    path = tmp_path / "results.csv"
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "member_id", "check_name", "M_kNm", "V_kN", "N_kN", "utilization", "compliant",
        ])
        w.writeheader()
        w.writerow({
            "member_id": member_id,
            "check_name": "ULS_bending",
            "M_kNm": str(m_kNm),
            "V_kN": "20",
            "N_kN": "0",
            "utilization": str(utilization),
            "compliant": "true",
        })
    return path


class TestResultsParser:
    def test_double_check_ok_when_match_exact(self, tmp_path: Path) -> None:
        """M analytique = 1.5 × 10 × 6² / 8 = 67.5 kNm. Logiciel = 67.5 → OK."""
        parser = SafResultsParser()
        csv_path = _write_csv_results(tmp_path, "B1", m_kNm=67.5)

        result = parser.parse_csv_results_and_check(
            csv_results_path=csv_path, model_data=MODEL_SIMPLE_BEAM,
        )

        assert len(result.member_checks) == 1
        assert result.member_checks[0].utilization_ratio == 0.8
        # Une anomalie INFO attendue (match OK)
        beam_anomalies = [a for a in result.anomalies if a.check_type == "beam_M_qL2_8"]
        assert len(beam_anomalies) == 1
        assert beam_anomalies[0].level == AnomalyLevel.INFO
        assert beam_anomalies[0].divergence_pct is not None
        assert beam_anomalies[0].divergence_pct < 1.0

    def test_double_check_warning_if_12_percent_divergence(self, tmp_path: Path) -> None:
        """Écart de 12% → WARNING (entre 10 et 15%)."""
        parser = SafResultsParser()
        # M analytique = 67.5 ; 67.5 × 1.12 = 75.6
        csv_path = _write_csv_results(tmp_path, "B1", m_kNm=75.6)

        result = parser.parse_csv_results_and_check(
            csv_results_path=csv_path, model_data=MODEL_SIMPLE_BEAM,
        )

        beam_anomalies = [a for a in result.anomalies if a.check_type == "beam_M_qL2_8"]
        assert len(beam_anomalies) == 1
        anomaly = beam_anomalies[0]
        assert anomaly.level == AnomalyLevel.WARNING
        assert DIVERGENCE_INFO_THRESHOLD_PCT < anomaly.divergence_pct <= DIVERGENCE_WARNING_THRESHOLD_PCT

    def test_double_check_flags_anomaly_if_divergence_over_15(self, tmp_path: Path) -> None:
        """Écart de 30% → ANOMALY."""
        parser = SafResultsParser()
        # M analytique = 67.5 ; 67.5 × 1.30 = 87.75
        csv_path = _write_csv_results(tmp_path, "B1", m_kNm=87.75)

        result = parser.parse_csv_results_and_check(
            csv_results_path=csv_path, model_data=MODEL_SIMPLE_BEAM,
        )

        beam_anomalies = [a for a in result.anomalies if a.check_type == "beam_M_qL2_8"]
        assert len(beam_anomalies) == 1
        anomaly = beam_anomalies[0]
        assert anomaly.level == AnomalyLevel.ANOMALY, \
            f"Divergence {anomaly.divergence_pct}% doit donner ANOMALY"
        assert anomaly.divergence_pct >= 15.0
        assert "ANOMALIE" in anomaly.message or "ANOMALY" in anomaly.message.upper()

    def test_compliant_false_if_anomaly_present(self, tmp_path: Path) -> None:
        parser = SafResultsParser()
        csv_path = _write_csv_results(tmp_path, "B1", m_kNm=87.75, utilization=0.9)

        result = parser.parse_csv_results_and_check(
            csv_results_path=csv_path, model_data=MODEL_SIMPLE_BEAM,
        )

        assert result.compliant is False, "Compliant doit être False s'il y a une ANOMALY"

    def test_utilization_over_1_triggers_non_compliant(self, tmp_path: Path) -> None:
        parser = SafResultsParser()
        csv_path = _write_csv_results(tmp_path, "B1", m_kNm=67.5, utilization=1.15)

        result = parser.parse_csv_results_and_check(
            csv_results_path=csv_path, model_data=MODEL_SIMPLE_BEAM,
        )

        assert result.max_utilization == 1.15
        assert result.compliant is False
        assert result.member_checks[0].compliant is False

    def test_saf_xlsx_parsing_extracts_results(self, tmp_path: Path) -> None:
        """Un SAF avec feuille 'Results' est parsé correctement."""
        # On crée un workbook minimal avec StructuralNodes + Members + Results + Loads
        wb = Workbook()
        wb.remove(wb.active)

        ws_nodes = wb.create_sheet("StructuralNodes")
        ws_nodes.append(["Name", "X_m", "Y_m", "Z_m"])
        ws_nodes.append(["N1", 0.0, 0.0, 3.0])
        ws_nodes.append(["N2", 6.0, 0.0, 3.0])

        ws_mem = wb.create_sheet("Structural1DMembers")
        ws_mem.append(["Name", "Type", "NodeStart", "NodeEnd", "CrossSection", "Material", "Length_m"])
        ws_mem.append(["B1", "beam", "N1", "N2", "POU_30x50", "C30/37", 6.0])

        ws_loads = wb.create_sheet("Loads")
        ws_loads.append(["Name", "LoadCase", "Target", "TargetType", "Type", "Direction", "Value"])
        ws_loads.append(["L1", "Q1", "B1", "member", "uniform_vertical", "-Z", 10.0])

        ws_res = wb.create_sheet("Results")
        ws_res.append(["Name", "My", "Vz", "Nx", "Utilization", "CheckName"])
        ws_res.append(["B1", 67.5, 30.0, 0.0, 0.85, "ULS_bending"])

        saf_path = tmp_path / "saf_enriched.xlsx"
        wb.save(saf_path)

        parser = SafResultsParser()
        result = parser.parse_and_check(
            saf_input=saf_path, saf_output=saf_path, model_data={},
        )

        assert len(result.member_checks) == 1
        assert result.member_checks[0].member_id == "B1"
        assert result.member_checks[0].utilization_ratio == 0.85
        assert result.raw_output["source"] == "saf_xlsx"

    def test_column_double_check(self, tmp_path: Path) -> None:
        """Poteau : somme des charges doit coller à N du logiciel à 15% près."""
        model = {
            "nodes": [
                {"id": "N1", "x": 0.0, "y": 0.0, "z": 0.0},
                {"id": "N3", "x": 0.0, "y": 0.0, "z": 3.0},
            ],
            "members": [
                {"id": "P1", "type": "column", "node_start": "N1", "node_end": "N3",
                 "section": "POT_30x30", "material": "C30/37"},
            ],
            "loads": [
                {"id": "L1", "case": "G1", "target": "P1", "target_type": "member",
                 "type": "point_vertical", "direction": "-Z",
                 "value_kN": 200.0, "category": "Permanent"},
            ],
        }
        # N analytique ELU = 200 × 1.35 = 270 kN
        import csv
        csv_path = tmp_path / "col_results.csv"
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=[
                "member_id", "check_name", "M_kNm", "V_kN", "N_kN", "utilization", "compliant",
            ])
            w.writeheader()
            w.writerow({
                "member_id": "P1", "check_name": "compression",
                "M_kNm": "0", "V_kN": "0", "N_kN": "272", "utilization": "0.6",
                "compliant": "true",
            })

        parser = SafResultsParser()
        result = parser.parse_csv_results_and_check(
            csv_results_path=csv_path, model_data=model,
        )

        col_anomalies = [a for a in result.anomalies if a.check_type == "column_N_sum"]
        assert len(col_anomalies) == 1
        # 270 vs 272 → ~0.7% → INFO
        assert col_anomalies[0].level == AnomalyLevel.INFO
        assert col_anomalies[0].divergence_pct < 5.0
