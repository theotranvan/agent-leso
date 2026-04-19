"""Tests unitaires pour SafGenerator (IFC → SAF xlsx buildingSMART)."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.connectors.structural import StructuralInputs
from app.connectors.structural.base import ConnectorError
from app.connectors.structural.saf_generator import (
    SIA_260_COMBINATIONS,
    STANDARD_LOAD_CASES,
    SafGenerator,
)


SAMPLE_MODEL_DATA = {
    "nodes": [
        {"id": "N1", "x": 0.0, "y": 0.0, "z": 0.0},
        {"id": "N2", "x": 6.0, "y": 0.0, "z": 0.0},
        {"id": "N3", "x": 0.0, "y": 0.0, "z": 3.0},
        {"id": "N4", "x": 6.0, "y": 0.0, "z": 3.0},
    ],
    "members": [
        {"id": "B1", "type": "beam", "node_start": "N3", "node_end": "N4",
         "section": "POU_30x50", "material": "C30/37"},
        {"id": "P1", "type": "column", "node_start": "N1", "node_end": "N3",
         "section": "POT_30x30", "material": "C30/37"},
        {"id": "P2", "type": "column", "node_start": "N2", "node_end": "N4",
         "section": "POT_30x30", "material": "C30/37"},
    ],
    "supports": [
        {"id": "S1", "node": "N1", "type": "fixed"},
        {"id": "S2", "node": "N2", "type": "fixed"},
    ],
}


class TestSafGenerator:
    def test_generates_valid_xlsx(self) -> None:
        gen = SafGenerator()
        inputs = StructuralInputs(ifc_path=None, model_data=SAMPLE_MODEL_DATA)
        xlsx_bytes = gen.generate_saf_bytes(inputs)

        assert len(xlsx_bytes) > 1000, "Le SAF doit faire plus de 1 KB"
        # Doit être un zip xlsx valide
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
        required_sheets = {
            "General", "Project", "Materials", "CrossSections",
            "StructuralNodes", "Structural1DMembers", "Supports",
            "LoadCases", "LoadCombinations", "Loads",
        }
        assert required_sheets.issubset(set(wb.sheetnames)), \
            f"Feuilles manquantes : {required_sheets - set(wb.sheetnames)}"

    def test_nodes_are_written_correctly(self) -> None:
        gen = SafGenerator()
        inputs = StructuralInputs(ifc_path=None, model_data=SAMPLE_MODEL_DATA)
        xlsx_bytes = gen.generate_saf_bytes(inputs)
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)

        rows = list(wb["StructuralNodes"].iter_rows(values_only=True))
        assert rows[0] == ("Name", "X_m", "Y_m", "Z_m")
        data_rows = rows[1:]
        assert len(data_rows) == 4, "4 nœuds attendus"

        # Vérifie N2 spécifiquement
        n2 = next(r for r in data_rows if r[0] == "N2")
        assert n2 == ("N2", 6.0, 0.0, 0.0)

    def test_members_include_length(self) -> None:
        gen = SafGenerator()
        inputs = StructuralInputs(ifc_path=None, model_data=SAMPLE_MODEL_DATA)
        xlsx_bytes = gen.generate_saf_bytes(inputs)
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)

        rows = list(wb["Structural1DMembers"].iter_rows(values_only=True))
        data_rows = rows[1:]
        b1_row = next(r for r in data_rows if r[0] == "B1")
        # B1 : N3(0,0,3) → N4(6,0,3) donc L = 6.0 m
        assert b1_row[-1] == 6.0, f"Longueur B1 attendue 6.0, obtenue {b1_row[-1]}"

        p1_row = next(r for r in data_rows if r[0] == "P1")
        # P1 : N1(0,0,0) → N3(0,0,3) donc L = 3.0 m
        assert p1_row[-1] == 3.0

    def test_all_sia_260_combinations_included(self) -> None:
        gen = SafGenerator()
        inputs = StructuralInputs(ifc_path=None, model_data=SAMPLE_MODEL_DATA)
        xlsx_bytes = gen.generate_saf_bytes(inputs)
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)

        rows = list(wb["LoadCombinations"].iter_rows(values_only=True))
        assert rows[0] == ("Name", "Description", "Type", "Factors_json")

        combo_ids = {r[0] for r in rows[1:] if r[0]}
        expected_ids = {c["id"] for c in SIA_260_COMBINATIONS}
        assert expected_ids == combo_ids, \
            f"Combinaisons SIA 260 manquantes : {expected_ids - combo_ids}"
        assert "ULS_STR_FUND" in combo_ids
        assert "ULS_ACC_EQU" in combo_ids
        assert "SLS_QP" in combo_ids

    def test_all_standard_load_cases_included(self) -> None:
        gen = SafGenerator()
        inputs = StructuralInputs(ifc_path=None, model_data=SAMPLE_MODEL_DATA)
        xlsx_bytes = gen.generate_saf_bytes(inputs)
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)

        rows = list(wb["LoadCases"].iter_rows(values_only=True))
        case_ids = {r[0] for r in rows[1:] if r[0]}
        expected = {c["id"] for c in STANDARD_LOAD_CASES}
        assert expected == case_ids, f"Cas manquants : {expected - case_ids}"

    def test_analyze_returns_saf_file_path(self, tmp_path: Path) -> None:
        gen = SafGenerator()
        inputs = StructuralInputs(ifc_path=None, model_data=SAMPLE_MODEL_DATA)
        result = gen.analyze(inputs)

        assert result.engine_used == "saf_generator"
        assert result.saf_file_path is not None
        assert result.saf_file_path.exists()
        assert result.saf_file_path.stat().st_size > 1000
        # Pas de calcul : compliant=False indéterminé, utilization=0
        assert result.max_utilization == 0.0

    def test_empty_model_raises(self) -> None:
        gen = SafGenerator()
        with pytest.raises(ConnectorError, match="Ni ifc_path ni model_data"):
            gen.analyze(StructuralInputs(ifc_path=None, model_data={}))
