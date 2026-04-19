"""Parser unifié de résultats structurels + double-check analytique.

## Principe du double-check analytique

Pour chaque élément analysé par le logiciel externe (Scia, Robot, etc.),
on recalcule la sollicitation principale avec une formule analytique simple :

- Poutre simplement appuyée avec charge uniforme : M_max = q·L² / 8
- Effort tranchant : V_max = q·L / 2
- Flèche : f_max = 5·q·L⁴ / (384·E·I)
- Poteau bi-articulé en compression : N = somme des charges

On compare à la valeur retournée par le logiciel. Si |écart| > 15%, on lève
un flag ANOMALY. Entre 10 et 15%, c'est un WARNING. Sous 10%, c'est OK.

Cette vérification indépendante ne remplace JAMAIS la validation d'un ingénieur
qualifié. Elle attrape les erreurs grossières (mauvaise longueur, charge
mal appliquée, modèle mal orienté).

## Formats supportés en entrée

1. SAF xlsx enrichi (feuille "Results" ou "Internal Forces")
2. CSV libre (colonnes : member_id, M_kNm, V_kN, N_kN, utilization...)
"""
from __future__ import annotations

import csv
import logging
import time
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.connectors.structural.base import (
    AnalysisResult,
    AnomalyLevel,
    ConnectorError,
    MemberCheck,
    StructuralAnomaly,
)

logger = logging.getLogger(__name__)

# Seuils de divergence analytique vs logiciel
DIVERGENCE_INFO_THRESHOLD_PCT: float = 10.0
DIVERGENCE_WARNING_THRESHOLD_PCT: float = 15.0

# Feuilles candidates pour les résultats dans un SAF enrichi
RESULT_SHEETS: tuple[str, ...] = (
    "Results",
    "InternalForces",
    "Internal1DForces",
    "MemberResults",
    "Utilizations",
    "DesignChecks",
)


class SafResultsParser:
    """Parse et double-check des résultats depuis SAF xlsx ou CSV libre."""

    def parse_and_check(
        self,
        saf_input: Path,
        saf_output: Path,
        model_data: dict[str, Any],
    ) -> AnalysisResult:
        """Parse SAF xlsx avec résultats + double-check."""
        start = time.monotonic()
        if not saf_output.exists():
            raise ConnectorError(f"SAF output introuvable : {saf_output}")

        try:
            wb = load_workbook(saf_output, data_only=True)
        except Exception as exc:
            raise ConnectorError(f"Lecture SAF échouée : {exc}") from exc

        software_results = self._extract_results_from_workbook(wb)
        if not software_results:
            raise ConnectorError(
                "Aucune feuille de résultats reconnue dans le SAF "
                f"(feuilles trouvées : {wb.sheetnames})"
            )

        # Lecture du modèle (nodes, members, loads) depuis les feuilles du SAF enrichi
        model = self._extract_model_from_workbook(wb)
        # Si le modèle SAF est vide, fallback sur model_data
        if not model.get("members") and model_data.get("members"):
            model = model_data

        result = self._run_double_check(model, software_results)
        result.computation_seconds = time.monotonic() - start
        result.raw_output["source"] = "saf_xlsx"
        result.raw_output["sheets_found"] = wb.sheetnames
        return result

    def parse_csv_results_and_check(
        self,
        csv_results_path: Path,
        model_data: dict[str, Any],
    ) -> AnalysisResult:
        """Parse CSV libre + double-check."""
        start = time.monotonic()
        if not csv_results_path.exists():
            raise ConnectorError(f"CSV résultats introuvable : {csv_results_path}")

        software_results: dict[str, dict[str, float]] = {}
        with open(csv_results_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                member_id = str(row.get("member_id") or "").strip()
                if not member_id:
                    continue
                software_results[member_id] = {
                    "M_kNm": self._safe_float(row.get("M_kNm")),
                    "V_kN": self._safe_float(row.get("V_kN")),
                    "N_kN": self._safe_float(row.get("N_kN")),
                    "utilization": self._safe_float(row.get("utilization"), default=0.0),
                    "check_name": str(row.get("check_name") or "unknown"),
                }

        if not software_results:
            raise ConnectorError("CSV résultats vide ou colonnes manquantes")

        result = self._run_double_check(model_data, software_results)
        result.computation_seconds = time.monotonic() - start
        result.raw_output["source"] = "csv"
        return result

    # ---------- core double-check ----------

    def _run_double_check(
        self,
        model: dict[str, Any],
        software_results: dict[str, dict[str, float]],
    ) -> AnalysisResult:
        member_checks: list[MemberCheck] = []
        anomalies: list[StructuralAnomaly] = []

        nodes_by_id = {n["id"]: n for n in model.get("nodes", [])}
        loads_by_target: dict[str, list[dict[str, Any]]] = {}
        for ld in model.get("loads", []):
            target = str(ld.get("target", ""))
            if target:
                loads_by_target.setdefault(target, []).append(ld)

        for member in model.get("members", []):
            member_id = str(member.get("id", ""))
            sw = software_results.get(member_id)
            if sw is None:
                # L'élément n'a pas de résultat logiciel - on enregistre comme info
                anomalies.append(StructuralAnomaly(
                    member_id=member_id,
                    check_type="missing_result",
                    level=AnomalyLevel.WARNING,
                    message="Aucun résultat logiciel pour cet élément",
                ))
                continue

            sw_utilization = sw.get("utilization", 0.0)
            member_checks.append(MemberCheck(
                member_id=member_id,
                check_name=sw.get("check_name", "unknown"),
                utilization_ratio=sw_utilization,
                compliant=sw_utilization <= 1.0,
                details={
                    "M_kNm_sw": sw.get("M_kNm"),
                    "V_kN_sw": sw.get("V_kN"),
                    "N_kN_sw": sw.get("N_kN"),
                },
            ))

            # Double-check analytique selon type
            member_type = str(member.get("type", "beam"))
            if member_type in ("beam", "girder"):
                anomaly = self._check_beam(member, nodes_by_id, loads_by_target, sw)
                if anomaly:
                    anomalies.append(anomaly)
            elif member_type in ("column", "post"):
                anomaly = self._check_column(member, loads_by_target, sw)
                if anomaly:
                    anomalies.append(anomaly)

        max_util = max((c.utilization_ratio for c in member_checks), default=0.0)
        has_anomaly = any(a.level == AnomalyLevel.ANOMALY for a in anomalies)
        compliant = max_util <= 1.0 and not has_anomaly

        return AnalysisResult(
            compliant=compliant,
            max_utilization=max_util,
            member_checks=member_checks,
            anomalies=anomalies,
            engine_used="saf_results_parser",
            warnings=[],
            raw_output={
                "nb_software_results": len(software_results),
                "nb_members_model": len(model.get("members", [])),
            },
        )

    def _check_beam(
        self,
        member: dict[str, Any],
        nodes_by_id: dict[str, dict[str, Any]],
        loads_by_target: dict[str, list[dict[str, Any]]],
        sw_result: dict[str, float],
    ) -> StructuralAnomaly | None:
        """Vérification analytique d'une poutre simplement appuyée : M_max = qL²/8."""
        member_id = str(member.get("id", ""))

        # Calcul de la longueur
        n_start = nodes_by_id.get(str(member.get("node_start", "")))
        n_end = nodes_by_id.get(str(member.get("node_end", "")))
        if not n_start or not n_end:
            return None

        L = (
            (n_end["x"] - n_start["x"]) ** 2
            + (n_end["y"] - n_start["y"]) ** 2
            + (n_end["z"] - n_start["z"]) ** 2
        ) ** 0.5
        if L <= 0:
            return None

        # Somme des charges uniformes verticales (kN/m) pondérées ELU 1.35/1.5
        q_total_uls = 0.0
        for load in loads_by_target.get(member_id, []):
            if str(load.get("type", "")) != "uniform_vertical":
                continue
            value = self._safe_float(load.get("value_kN_m"))
            # Pondération indicative selon catégorie
            category = (load.get("category") or "Variable").lower()
            if "permanent" in category:
                factor = 1.35
            elif "accident" in category:
                factor = 1.0
            else:
                factor = 1.5
            q_total_uls += abs(value) * factor

        if q_total_uls <= 0:
            # Pas de charge explicite - impossible de comparer
            return None

        # M_max analytique = q·L² / 8
        m_analytical = q_total_uls * L * L / 8.0
        m_software = abs(self._safe_float(sw_result.get("M_kNm")))
        if m_software <= 0:
            return None

        divergence_pct = abs(m_analytical - m_software) / m_analytical * 100.0

        if divergence_pct <= DIVERGENCE_INFO_THRESHOLD_PCT:
            level = AnomalyLevel.INFO
            msg = f"Poutre {member_id} : M analytique {m_analytical:.2f} kNm vs logiciel {m_software:.2f} kNm (écart {divergence_pct:.1f}%) - OK"
        elif divergence_pct <= DIVERGENCE_WARNING_THRESHOLD_PCT:
            level = AnomalyLevel.WARNING
            msg = f"Poutre {member_id} : écart M de {divergence_pct:.1f}% (analytique {m_analytical:.2f} vs logiciel {m_software:.2f}) - à vérifier"
        else:
            level = AnomalyLevel.ANOMALY
            msg = (
                f"Poutre {member_id} : écart M de {divergence_pct:.1f}% > 15% "
                f"(analytique {m_analytical:.2f} kNm vs logiciel {m_software:.2f} kNm) - ANOMALIE à investiguer"
            )

        return StructuralAnomaly(
            member_id=member_id,
            check_type="beam_M_qL2_8",
            level=level,
            message=msg,
            analytical_value=round(m_analytical, 3),
            software_value=round(m_software, 3),
            divergence_pct=round(divergence_pct, 2),
        )

    def _check_column(
        self,
        member: dict[str, Any],
        loads_by_target: dict[str, list[dict[str, Any]]],
        sw_result: dict[str, float],
    ) -> StructuralAnomaly | None:
        """Poteau bi-articulé en compression : N = somme charges ponctuelles verticales."""
        member_id = str(member.get("id", ""))

        n_total = 0.0
        for load in loads_by_target.get(member_id, []):
            if str(load.get("type", "")) != "point_vertical":
                continue
            value = self._safe_float(load.get("value_kN"))
            category = (load.get("category") or "Variable").lower()
            if "permanent" in category:
                factor = 1.35
            elif "accident" in category:
                factor = 1.0
            else:
                factor = 1.5
            n_total += abs(value) * factor

        if n_total <= 0:
            return None

        n_software = abs(self._safe_float(sw_result.get("N_kN")))
        if n_software <= 0:
            return None

        divergence_pct = abs(n_total - n_software) / n_total * 100.0

        if divergence_pct <= DIVERGENCE_INFO_THRESHOLD_PCT:
            level = AnomalyLevel.INFO
            msg = f"Poteau {member_id} : N analytique {n_total:.2f} kN vs logiciel {n_software:.2f} kN ({divergence_pct:.1f}%) - OK"
        elif divergence_pct <= DIVERGENCE_WARNING_THRESHOLD_PCT:
            level = AnomalyLevel.WARNING
            msg = f"Poteau {member_id} : écart N de {divergence_pct:.1f}% - à vérifier"
        else:
            level = AnomalyLevel.ANOMALY
            msg = (
                f"Poteau {member_id} : écart N de {divergence_pct:.1f}% > 15% "
                f"(analytique {n_total:.2f} vs logiciel {n_software:.2f}) - ANOMALIE"
            )

        return StructuralAnomaly(
            member_id=member_id,
            check_type="column_N_sum",
            level=level,
            message=msg,
            analytical_value=round(n_total, 3),
            software_value=round(n_software, 3),
            divergence_pct=round(divergence_pct, 2),
        )

    # ---------- extraction depuis workbook ----------

    def _extract_results_from_workbook(self, wb: Any) -> dict[str, dict[str, float]]:
        """Trouve la feuille de résultats et extrait les lignes."""
        results: dict[str, dict[str, float]] = {}
        target_sheet = None
        for sheet_name in RESULT_SHEETS:
            if sheet_name in wb.sheetnames:
                target_sheet = wb[sheet_name]
                break
        if target_sheet is None:
            return results

        rows = list(target_sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return results

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

        def _find_col(candidates: list[str]) -> int | None:
            for idx, h in enumerate(headers):
                for c in candidates:
                    if c.lower() in h.lower():
                        return idx
            return None

        col_id = _find_col(["Name", "Member", "MemberID", "member_id", "Id"])
        col_m = _find_col(["My", "M_y", "Moment", "M_kNm", "Mmax"])
        col_v = _find_col(["Vz", "V_z", "Shear", "V_kN", "Vmax"])
        col_n = _find_col(["Nx", "N_x", "Normal", "N_kN"])
        col_util = _find_col(["Utilization", "Ratio", "UnityCheck", "Util"])
        col_check = _find_col(["Check", "CheckName", "check_name"])

        if col_id is None:
            return results

        for row in rows[1:]:
            if row[col_id] is None:
                continue
            member_id = str(row[col_id]).strip()
            if not member_id:
                continue
            results[member_id] = {
                "M_kNm": self._safe_float(row[col_m] if col_m is not None else None),
                "V_kN": self._safe_float(row[col_v] if col_v is not None else None),
                "N_kN": self._safe_float(row[col_n] if col_n is not None else None),
                "utilization": self._safe_float(row[col_util] if col_util is not None else None, default=0.0),
                "check_name": str(row[col_check]) if col_check is not None and row[col_check] else "unknown",
            }

        return results

    def _extract_model_from_workbook(self, wb: Any) -> dict[str, Any]:
        """Extrait nodes + members + loads depuis les feuilles SAF standards."""
        model: dict[str, Any] = {"nodes": [], "members": [], "loads": []}

        if "StructuralNodes" in wb.sheetnames:
            ws = wb["StructuralNodes"]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) >= 2:
                for row in rows[1:]:
                    if row[0] is None:
                        continue
                    model["nodes"].append({
                        "id": str(row[0]),
                        "x": self._safe_float(row[1]),
                        "y": self._safe_float(row[2]),
                        "z": self._safe_float(row[3]),
                    })

        if "Structural1DMembers" in wb.sheetnames:
            ws = wb["Structural1DMembers"]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) >= 2:
                for row in rows[1:]:
                    if row[0] is None:
                        continue
                    model["members"].append({
                        "id": str(row[0]),
                        "type": str(row[1]) if len(row) > 1 and row[1] else "beam",
                        "node_start": str(row[2]) if len(row) > 2 and row[2] else "",
                        "node_end": str(row[3]) if len(row) > 3 and row[3] else "",
                        "section": str(row[4]) if len(row) > 4 and row[4] else "",
                        "material": str(row[5]) if len(row) > 5 and row[5] else "",
                    })

        if "Loads" in wb.sheetnames:
            ws = wb["Loads"]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) >= 2:
                for row in rows[1:]:
                    if not row or row[0] is None:
                        continue
                    model["loads"].append({
                        "id": str(row[0]),
                        "case": str(row[1]) if len(row) > 1 and row[1] else "",
                        "target": str(row[2]) if len(row) > 2 and row[2] else "",
                        "target_type": str(row[3]) if len(row) > 3 and row[3] else "member",
                        "type": str(row[4]) if len(row) > 4 and row[4] else "uniform_vertical",
                        "direction": str(row[5]) if len(row) > 5 and row[5] else "-Z",
                        "value_kN_m": self._safe_float(row[6] if len(row) > 6 else None),
                        "value_kN": self._safe_float(row[6] if len(row) > 6 else None),
                    })

        return model

    @staticmethod
    def _safe_float(v: Any, default: float = 0.0) -> float:
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        try:
            s = str(v).strip().replace(",", ".").replace("'", "")
            return float(s) if s else default
        except (ValueError, TypeError):
            return default
