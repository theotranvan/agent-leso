"""Générateur SAF (Structural Analysis Format buildingSMART) depuis un IFC.

SAF est un format d'échange xlsx normé par buildingSMART International,
supporté nativement par Scia Engineer, RFEM/RSTAB (Dlubal), et partiellement
par d'autres (Cedrus, Advance Design).

Pipeline :
1. Ouvrir l'IFC avec ifcopenshell
2. Extraire : IfcBeam, IfcColumn, IfcSlab, IfcWall (structural role)
3. Construire le modèle de nœuds / éléments 1D
4. Générer les cas de charge standards SIA 261 (G, Q, neige, vent, séisme)
5. Générer les combinaisons SIA 260 (ELU fondamentale + ELS caractéristique/fréquente/qp)
6. Écrire le xlsx avec les feuilles SAF officielles

Limitations V3 :
- Géométrie linéaire (1D), pas de surfaces structurelles 2D
- Sections rectangulaires/profilés standards uniquement
- Séisme forfaitaire (pas de spectres, pas de masse modale)
- Sol = appuis encastrés/articulés (pas d'interaction sol-structure)
"""
from __future__ import annotations

import logging
import tempfile
import time
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.connectors.structural.base import (
    AnalysisResult,
    ConnectorError,
    StructuralConnector,
    StructuralInputs,
)

logger = logging.getLogger(__name__)

SAF_VERSION = "2.1.0"

# Combinaisons SIA 260 standards
SIA_260_COMBINATIONS: list[dict[str, Any]] = [
    {
        "id": "ULS_STR_FUND",
        "name": "ELU STR fondamentale : 1.35·G + 1.5·Q + 1.5·ψ0,S·S + 1.5·ψ0,W·W",
        "type": "linear_add",
        "factors": [
            {"case": "G1", "factor": 1.35},
            {"case": "G2", "factor": 1.35},
            {"case": "Q1", "factor": 1.5},
            {"case": "QS", "factor": 0.75},
            {"case": "QW", "factor": 0.9},
        ],
    },
    {
        "id": "ULS_ACC_EQU",
        "name": "ELU accidentelle (sismique) : G + A_E + ψ2·Q",
        "type": "linear_add",
        "factors": [
            {"case": "G1", "factor": 1.0},
            {"case": "G2", "factor": 1.0},
            {"case": "Q1", "factor": 0.3},
            {"case": "AE", "factor": 1.0},
        ],
    },
    {
        "id": "SLS_CHAR",
        "name": "ELS caractéristique : G + Q + ψ0,S·S",
        "type": "linear_add",
        "factors": [
            {"case": "G1", "factor": 1.0},
            {"case": "G2", "factor": 1.0},
            {"case": "Q1", "factor": 1.0},
            {"case": "QS", "factor": 0.5},
        ],
    },
    {
        "id": "SLS_FREQ",
        "name": "ELS fréquente : G + ψ1·Q",
        "type": "linear_add",
        "factors": [
            {"case": "G1", "factor": 1.0},
            {"case": "G2", "factor": 1.0},
            {"case": "Q1", "factor": 0.5},
        ],
    },
    {
        "id": "SLS_QP",
        "name": "ELS quasi-permanente : G + ψ2·Q",
        "type": "linear_add",
        "factors": [
            {"case": "G1", "factor": 1.0},
            {"case": "G2", "factor": 1.0},
            {"case": "Q1", "factor": 0.3},
        ],
    },
]


STANDARD_LOAD_CASES: list[dict[str, str]] = [
    {"id": "G1", "name": "Poids propre", "category": "Permanent"},
    {"id": "G2", "name": "Charges permanentes revêtements", "category": "Permanent"},
    {"id": "Q1", "name": "Charge d'exploitation (catégorie A)", "category": "Variable"},
    {"id": "QS", "name": "Neige", "category": "Variable"},
    {"id": "QW", "name": "Vent", "category": "Variable"},
    {"id": "AE", "name": "Action sismique", "category": "Accidentel"},
]


MATERIAL_PROPERTIES: dict[str, dict[str, Any]] = {
    "C25/30": {"class": "Concrete", "density_kg_m3": 2500, "E_GPa": 31, "fck_MPa": 25},
    "C30/37": {"class": "Concrete", "density_kg_m3": 2500, "E_GPa": 33, "fck_MPa": 30},
    "C35/45": {"class": "Concrete", "density_kg_m3": 2500, "E_GPa": 34, "fck_MPa": 35},
    "S235": {"class": "Steel", "density_kg_m3": 7850, "E_GPa": 210, "fy_MPa": 235},
    "S355": {"class": "Steel", "density_kg_m3": 7850, "E_GPa": 210, "fy_MPa": 355},
    "GL24h": {"class": "Timber", "density_kg_m3": 420, "E_GPa": 11.5, "fmk_MPa": 24},
    "C24": {"class": "Timber", "density_kg_m3": 420, "E_GPa": 11, "fmk_MPa": 24},
    "B500B": {"class": "Reinforcement", "E_GPa": 200, "fyk_MPa": 500},
}


@dataclass
class SafNode:
    id: str
    x: float
    y: float
    z: float


@dataclass
class SafMember:
    id: str
    node_start: str
    node_end: str
    section: str
    material: str
    member_type: str = "beam"


@dataclass
class SafModel:
    nodes: list[SafNode] = field(default_factory=list)
    members: list[SafMember] = field(default_factory=list)
    supports: list[dict[str, Any]] = field(default_factory=list)


class SafGenerator(StructuralConnector):
    """Génère un fichier SAF xlsx depuis un IFC ou un model_data JSON."""

    name = "saf_generator"

    def validate_inputs(self, inputs: StructuralInputs) -> list[str]:
        warnings: list[str] = []
        if inputs.ifc_path is not None:
            if not inputs.ifc_path.exists():
                raise ConnectorError(f"IFC introuvable : {inputs.ifc_path}")
        elif not inputs.model_data:
            raise ConnectorError("Ni ifc_path ni model_data fournis")
        if inputs.referentiel not in {"sia", "eurocode"}:
            warnings.append(f"Référentiel inconnu : {inputs.referentiel} - SIA utilisé")
        return warnings

    def analyze(self, inputs: StructuralInputs) -> AnalysisResult:
        """Pour SafGenerator, 'analyze' génère le SAF sans calcul - calcul = logiciel externe."""
        start = time.monotonic()
        warnings = self.validate_inputs(inputs)

        model = self._build_model(inputs, warnings)
        if not model.members:
            raise ConnectorError("Aucun élément structurel trouvé")

        xlsx_bytes = self._build_saf_xlsx(model, inputs)
        tmp_dir = Path(tempfile.gettempdir())
        tmp_path = tmp_dir / f"saf_{int(time.time())}.xlsx"
        tmp_path.write_bytes(xlsx_bytes)

        elapsed = time.monotonic() - start
        logger.info(
            "SAF généré : %d nœuds, %d éléments, fichier %.1f KB en %.2fs",
            len(model.nodes), len(model.members), len(xlsx_bytes) / 1024, elapsed,
        )

        # Pas de calcul ni vérif : le SAF est en attente d'enrichissement logiciel
        return AnalysisResult(
            compliant=False,  # indéterminé tant que pas de calcul
            max_utilization=0.0,
            member_checks=[],
            anomalies=[],
            engine_used=self.name,
            computation_seconds=elapsed,
            warnings=warnings + [
                "SAF généré sans calcul - à enrichir via Scia/RFEM puis réimporter pour vérifications",
            ],
            saf_file_path=tmp_path,
            raw_output={
                "nb_nodes": len(model.nodes),
                "nb_members": len(model.members),
                "nb_supports": len(model.supports),
                "saf_bytes_length": len(xlsx_bytes),
            },
        )

    def generate_saf_bytes(self, inputs: StructuralInputs) -> bytes:
        """Expose la génération brute sans créer d'AnalysisResult."""
        warnings: list[str] = []
        model = self._build_model(inputs, warnings)
        return self._build_saf_xlsx(model, inputs)

    # ---------- construction du modèle ----------

    def _build_model(self, inputs: StructuralInputs, warnings: list[str]) -> SafModel:
        if inputs.ifc_path is not None:
            return self._build_from_ifc(inputs.ifc_path, warnings)
        return self._build_from_data(inputs.model_data, warnings)

    def _build_from_ifc(self, ifc_path: Path, warnings: list[str]) -> SafModel:
        try:
            import ifcopenshell
        except ImportError as exc:
            raise ConnectorError("ifcopenshell non installé") from exc

        try:
            ifc = ifcopenshell.open(str(ifc_path))
        except Exception as exc:
            raise ConnectorError(f"Lecture IFC échouée : {exc}") from exc

        model = SafModel()
        node_counter = 1

        for ifc_class, member_type, section_fallback in [
            ("IfcBeam", "beam", "POU_30x50"),
            ("IfcColumn", "column", "POT_30x30"),
        ]:
            try:
                elements = ifc.by_type(ifc_class)
            except Exception as exc:
                warnings.append(f"Extraction {ifc_class} échouée : {exc}")
                continue

            for elem in elements:
                start_pt, end_pt = self._extract_axis(elem)
                if start_pt is None or end_pt is None:
                    warnings.append(f"{ifc_class} {elem.id()} sans axe : ignoré")
                    continue
                n1_id = f"N{node_counter}"
                n2_id = f"N{node_counter + 1}"
                node_counter += 2
                model.nodes.append(SafNode(n1_id, *start_pt))
                model.nodes.append(SafNode(n2_id, *end_pt))

                section = self._extract_section(elem) or section_fallback
                material = self._extract_material(elem) or "C30/37"

                model.members.append(SafMember(
                    id=elem.GlobalId or f"{ifc_class}_{elem.id()}",
                    node_start=n1_id,
                    node_end=n2_id,
                    section=section,
                    material=material,
                    member_type=member_type,
                ))

        # Appuis : nœuds en Z=0 considérés comme appuis encastrés par défaut
        for node in model.nodes:
            if abs(node.z) < 0.01:
                model.supports.append({
                    "id": f"SUP_{node.id}",
                    "node": node.id,
                    "type": "fixed",
                })

        return model

    def _build_from_data(self, data: dict[str, Any], warnings: list[str]) -> SafModel:
        model = SafModel()
        for n in data.get("nodes", []):
            model.nodes.append(SafNode(
                id=n["id"],
                x=float(n["x"]),
                y=float(n["y"]),
                z=float(n["z"]),
            ))
        for m in data.get("members", []):
            model.members.append(SafMember(
                id=m["id"],
                node_start=m["node_start"],
                node_end=m["node_end"],
                section=m.get("section", "POU_30x50"),
                material=m.get("material", "C30/37"),
                member_type=m.get("type", "beam"),
            ))
        for s in data.get("supports", []):
            model.supports.append(s)
        return model

    @staticmethod
    def _extract_axis(element: Any) -> tuple[tuple[float, float, float] | None, tuple[float, float, float] | None]:
        """Extrait points début/fin d'un IfcBeam ou IfcColumn.

        Approche pragmatique : lit l'ObjectPlacement + l'axe local dans Representation.
        Si échec, retourne (None, None) et l'élément est ignoré.
        """
        try:
            placement = element.ObjectPlacement
            if placement is None:
                return None, None
            rel_placement = getattr(placement, "RelativePlacement", None)
            if rel_placement is None:
                return None, None
            location = getattr(rel_placement, "Location", None)
            if location is None or not getattr(location, "Coordinates", None):
                return None, None
            coords = list(location.Coordinates)
            if len(coords) < 3:
                coords = coords + [0.0] * (3 - len(coords))

            # Longueur par défaut (3m vertical pour colonne, horizontal pour poutre)
            if element.is_a() == "IfcColumn":
                end = (coords[0], coords[1], coords[2] + 3.0)
            else:
                # Tente de lire la longueur via Qto
                import ifcopenshell.util.element as util
                psets = util.get_psets(element) or {}
                length = None
                for pset_data in psets.values():
                    if isinstance(pset_data, dict):
                        length = pset_data.get("Length") or length
                if not isinstance(length, (int, float)) or length <= 0:
                    length = 5.0
                end = (coords[0] + float(length), coords[1], coords[2])
            return tuple(coords[:3]), end
        except Exception:
            return None, None

    @staticmethod
    def _extract_section(element: Any) -> str | None:
        """Tente d'extraire une section depuis Pset_BeamCommon / Pset_ColumnCommon."""
        try:
            import ifcopenshell.util.element as util
            psets = util.get_psets(element) or {}
            for pset_name in ("Pset_BeamCommon", "Pset_ColumnCommon", "BETAgent_MemberData"):
                p = psets.get(pset_name)
                if isinstance(p, dict):
                    for key in ("Reference", "Section", "CrossSection"):
                        v = p.get(key)
                        if isinstance(v, str) and v.strip():
                            return v.strip()
        except Exception:
            pass
        return None

    @staticmethod
    def _extract_material(element: Any) -> str | None:
        try:
            import ifcopenshell.util.element as util
            mats = util.get_material(element) or None
            if mats is None:
                return None
            if hasattr(mats, "Name") and mats.Name:
                return mats.Name
            if hasattr(mats, "Materials") and mats.Materials:
                first = mats.Materials[0]
                return getattr(first, "Name", None)
        except Exception:
            pass
        return None

    # ---------- construction du xlsx ----------

    def _build_saf_xlsx(self, model: SafModel, inputs: StructuralInputs) -> bytes:
        wb = Workbook()
        # Retire la feuille par défaut
        wb.remove(wb.active)

        self._sheet_general(wb)
        self._sheet_project(wb, inputs)
        self._sheet_materials(wb, model)
        self._sheet_sections(wb, model)
        self._sheet_nodes(wb, model)
        self._sheet_members(wb, model)
        self._sheet_supports(wb, model)
        self._sheet_load_cases(wb)
        self._sheet_combinations(wb)
        self._sheet_loads(wb)  # vide initialement, rempli par l'ingénieur

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _header_style(cell: Any) -> None:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _sheet_general(self, wb: Workbook) -> None:
        ws = wb.create_sheet("General")
        import time as _time
        rows = [
            ["SAF Version", SAF_VERSION],
            ["Generator", "BET Agent V3 - SafGenerator"],
            ["Generated At", _time.strftime("%Y-%m-%dT%H:%M:%SZ", _time.gmtime())],
            ["Units Length", "m"],
            ["Units Force", "kN"],
            ["Units Mass", "t"],
            ["Units Temperature", "C"],
        ]
        for r in rows:
            ws.append(r)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 40

    def _sheet_project(self, wb: Workbook, inputs: StructuralInputs) -> None:
        ws = wb.create_sheet("Project")
        rows = [
            ["Referentiel", inputs.referentiel],
            ["Standard", "SIA 260-267" if inputs.referentiel == "sia" else "Eurocodes"],
            ["ExposureClass", inputs.exposure_class],
            ["ConsequenceClass", inputs.consequence_class],
            ["SeismicZone", inputs.seismic_zone],
            ["MaterialDefault", inputs.material_default],
        ]
        for r in rows:
            ws.append(r)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 25

    def _sheet_materials(self, wb: Workbook, model: SafModel) -> None:
        ws = wb.create_sheet("Materials")
        headers = ["Name", "Class", "Density_kg_m3", "E_GPa", "fck_MPa", "fy_MPa", "fmk_MPa", "fyk_MPa"]
        ws.append(headers)
        for cell in ws[1]:
            self._header_style(cell)

        used = {m.material for m in model.members} | {"C30/37", "S355"}
        for name in sorted(used):
            props = MATERIAL_PROPERTIES.get(name, {"class": "Unknown"})
            ws.append([
                name,
                props.get("class"),
                props.get("density_kg_m3"),
                props.get("E_GPa"),
                props.get("fck_MPa"),
                props.get("fy_MPa"),
                props.get("fmk_MPa"),
                props.get("fyk_MPa"),
            ])

    def _sheet_sections(self, wb: Workbook, model: SafModel) -> None:
        ws = wb.create_sheet("CrossSections")
        headers = ["Name", "Shape", "h_mm", "b_mm", "Material", "Usage"]
        ws.append(headers)
        for cell in ws[1]:
            self._header_style(cell)

        defaults = {
            "POT_25x25": ("Rectangular", 250, 250, "Poteau BA"),
            "POT_30x30": ("Rectangular", 300, 300, "Poteau BA"),
            "POU_30x50": ("Rectangular", 500, 300, "Poutre BA"),
            "HEB200": ("HEB", 200, 200, "Profilé acier"),
            "IPE200": ("IPE", 200, 100, "Profilé acier"),
            "DALLE_20cm": ("Plate", 200, None, "Dalle BA"),
        }
        used_sections = {m.section for m in model.members} | set(defaults.keys())
        for name in sorted(used_sections):
            shape_data = defaults.get(name, ("Rectangular", 300, 300, ""))
            ws.append([name, shape_data[0], shape_data[1], shape_data[2], "", shape_data[3]])

    def _sheet_nodes(self, wb: Workbook, model: SafModel) -> None:
        ws = wb.create_sheet("StructuralNodes")
        ws.append(["Name", "X_m", "Y_m", "Z_m"])
        for cell in ws[1]:
            self._header_style(cell)
        for n in model.nodes:
            ws.append([n.id, n.x, n.y, n.z])

    def _sheet_members(self, wb: Workbook, model: SafModel) -> None:
        ws = wb.create_sheet("Structural1DMembers")
        ws.append(["Name", "Type", "NodeStart", "NodeEnd", "CrossSection", "Material", "Length_m"])
        for cell in ws[1]:
            self._header_style(cell)

        nodes_by_id = {n.id: n for n in model.nodes}
        for m in model.members:
            ns = nodes_by_id.get(m.node_start)
            ne = nodes_by_id.get(m.node_end)
            length = None
            if ns and ne:
                length = round(((ne.x - ns.x) ** 2 + (ne.y - ns.y) ** 2 + (ne.z - ns.z) ** 2) ** 0.5, 3)
            ws.append([m.id, m.member_type, m.node_start, m.node_end, m.section, m.material, length])

    def _sheet_supports(self, wb: Workbook, model: SafModel) -> None:
        ws = wb.create_sheet("Supports")
        ws.append(["Name", "Node", "Type", "RX", "RY", "RZ", "MX", "MY", "MZ"])
        for cell in ws[1]:
            self._header_style(cell)
        for s in model.supports:
            t = s.get("type", "fixed")
            restraints = {
                "fixed": (True, True, True, True, True, True),
                "pinned": (True, True, True, False, False, False),
                "roller": (False, False, True, False, False, False),
            }.get(t, (True, True, True, False, False, False))
            ws.append([s.get("id", ""), s.get("node", ""), t, *restraints])

    def _sheet_load_cases(self, wb: Workbook) -> None:
        ws = wb.create_sheet("LoadCases")
        ws.append(["Name", "Description", "ActionType"])
        for cell in ws[1]:
            self._header_style(cell)
        for lc in STANDARD_LOAD_CASES:
            ws.append([lc["id"], lc["name"], lc["category"]])

    def _sheet_combinations(self, wb: Workbook) -> None:
        import json as _json

        ws = wb.create_sheet("LoadCombinations")
        ws.append(["Name", "Description", "Type", "Factors_json"])
        for cell in ws[1]:
            self._header_style(cell)
        for combo in SIA_260_COMBINATIONS:
            ws.append([
                combo["id"],
                combo["name"],
                combo["type"],
                _json.dumps(combo["factors"], ensure_ascii=False),
            ])
        ws.column_dimensions["D"].width = 80

    def _sheet_loads(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Loads")
        ws.append(["Name", "LoadCase", "Target", "TargetType", "Type", "Direction", "Value_kN_or_kNm"])
        for cell in ws[1]:
            self._header_style(cell)
        # Vide - à remplir par l'ingénieur dans le logiciel de calcul
