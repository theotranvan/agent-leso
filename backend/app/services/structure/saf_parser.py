"""Parser SAF - import des résultats depuis un xlsx SAF enrichi par le logiciel de calcul.

Quand l'ingénieur relance ses calculs dans Scia/RFEM, le SAF exporté contient des feuilles
de résultats. On les lit ici et on extrait sollicitations + réactions + taux de travail.
"""
import logging
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_saf_results(xlsx_bytes: bytes) -> dict:
    """Extrait les résultats d'un SAF enrichi.

    Cherche les feuilles courantes :
    - InternalForces / MemberResults
    - Reactions / SupportReactions
    - Deformations
    - UtilizationRatios / Utilizations

    Retourne un dict structuré pour stockage + analyse.
    """
    try:
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        logger.error(f"Impossible d'ouvrir le SAF : {e}")
        return {"error": str(e)}

    results: dict[str, Any] = {
        "sheets_found": wb.sheetnames,
        "internal_forces": [],
        "reactions": [],
        "deformations": [],
        "utilizations": [],
        "warnings": [],
    }

    # Sheets candidates
    candidates_if = ["InternalForces", "MemberResults", "Internal1DForces", "Forces"]
    candidates_reac = ["Reactions", "SupportReactions"]
    candidates_def = ["Deformations", "Displacements", "NodeDeformations"]
    candidates_util = ["UtilizationRatios", "Utilizations", "Design", "DesignChecks"]

    def _find_sheet(candidates):
        for c in candidates:
            if c in wb.sheetnames:
                return wb[c]
        return None

    # Forces internes
    ws = _find_sheet(candidates_if)
    if ws:
        results["internal_forces"] = _read_sheet_as_dicts(ws)

    ws = _find_sheet(candidates_reac)
    if ws:
        results["reactions"] = _read_sheet_as_dicts(ws)

    ws = _find_sheet(candidates_def)
    if ws:
        results["deformations"] = _read_sheet_as_dicts(ws)

    ws = _find_sheet(candidates_util)
    if ws:
        results["utilizations"] = _read_sheet_as_dicts(ws)

    # Détection anomalies basiques
    max_util = 0.0
    for u in results["utilizations"]:
        for key in ("Utilization", "Ratio", "Util", "UnityCheck"):
            v = u.get(key)
            if isinstance(v, (int, float)):
                max_util = max(max_util, float(v))
    if max_util > 1.0:
        results["warnings"].append(
            f"Taux de travail maximal = {max_util:.2f} (> 1.0) - section(s) insuffisante(s)"
        )
    elif max_util > 0.95:
        results["warnings"].append(
            f"Taux de travail maximal = {max_util:.2f} (> 0.95) - marge faible, à vérifier"
        )

    results["max_utilization"] = max_util
    return results


def _read_sheet_as_dicts(ws) -> list[dict]:
    """Lit une feuille xlsx avec 1ère ligne = headers, retourne liste de dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        data.append(d)
    return data
