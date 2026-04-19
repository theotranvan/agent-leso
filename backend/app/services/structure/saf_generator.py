"""Générateur SAF - Structural Analysis Format (buildingSMART).

SAF est un format d'échange ouvert pour modèles de calcul structurel,
supporté par Scia Engineer, RFEM/RSTAB (Dlubal), partiellement par d'autres.

Format SAF v2.10 : tableur (xlsx) structuré avec feuilles :
- General
- Project
- Materials
- CrossSections
- StructuralNodes
- Structural1DMembers
- StructuralSurfaces
- Supports
- LoadCases / LoadGroups / LoadCombinations
- Loads

Ce module génère un xlsx conforme au schema minimal SAF.
Spec complète : https://www.saf.guide/
"""
import logging
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

SAF_VERSION = "2.1.0"


# Matériaux SIA par défaut
MATERIALS_SIA = {
    "C25/30": {"class": "Concrete", "grade": "C25/30", "density_kg_m3": 2500, "E_GPa": 31, "fck_MPa": 25},
    "C30/37": {"class": "Concrete", "grade": "C30/37", "density_kg_m3": 2500, "E_GPa": 33, "fck_MPa": 30},
    "C35/45": {"class": "Concrete", "grade": "C35/45", "density_kg_m3": 2500, "E_GPa": 34, "fck_MPa": 35},
    "S235": {"class": "Steel", "grade": "S235", "density_kg_m3": 7850, "E_GPa": 210, "fy_MPa": 235},
    "S355": {"class": "Steel", "grade": "S355", "density_kg_m3": 7850, "E_GPa": 210, "fy_MPa": 355},
    "GL24h": {"class": "Timber", "grade": "GL24h", "density_kg_m3": 420, "E_GPa": 11.5, "fmk_MPa": 24},
    "C24": {"class": "Timber", "grade": "C24", "density_kg_m3": 420, "E_GPa": 11, "fmk_MPa": 24},
    "B500B": {"class": "Reinforcement", "grade": "B500B", "fyk_MPa": 500, "E_GPa": 200},
}


# Sections transversales courantes
CROSS_SECTIONS_DEFAULTS = {
    "POT_25x25": {"shape": "Rectangular", "h_mm": 250, "b_mm": 250, "usage": "Poteau BA"},
    "POT_30x30": {"shape": "Rectangular", "h_mm": 300, "b_mm": 300, "usage": "Poteau BA"},
    "POU_30x50": {"shape": "Rectangular", "h_mm": 500, "b_mm": 300, "usage": "Poutre BA"},
    "HEB200": {"shape": "HEB", "h_mm": 200, "usage": "Profilé acier"},
    "IPE200": {"shape": "IPE", "h_mm": 200, "usage": "Profilé acier"},
    "DALLE_18cm": {"shape": "Plate", "thickness_mm": 180, "usage": "Dalle BA"},
    "DALLE_20cm": {"shape": "Plate", "thickness_mm": 200, "usage": "Dalle BA"},
    "DALLE_22cm": {"shape": "Plate", "thickness_mm": 220, "usage": "Dalle BA"},
    "DALLE_25cm": {"shape": "Plate", "thickness_mm": 250, "usage": "Dalle BA"},
}


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_saf_xlsx(structural_model: dict) -> bytes:
    """Génère un fichier SAF xlsx à partir d'un modèle structurel.

    structural_model attendu :
    {
      "project": {"name": "...", "referentiel": "sia"},
      "nodes": [{"id": "N1", "x": 0, "y": 0, "z": 0}, ...],
      "members": [{"id": "M1", "node_start": "N1", "node_end": "N2", "section": "HEB200", "material": "S355", "type": "beam"}, ...],
      "supports": [{"id": "S1", "node": "N1", "type": "fixed|pinned|roller", "restraint": "UX,UY,UZ"}, ...],
      "load_cases": [{"id": "LC1", "name": "Permanente", "category": "G"}, ...],
      "loads": [{"id": "L1", "case": "LC1", "target": "M1", "type": "uniform_vertical", "value_kN_m": -5.0}, ...],
      "combinations": [{"id": "ULS1", "name": "1.35G + 1.5Q", "type": "linear_add",
                        "factors": [{"case": "LC1", "factor": 1.35}, {"case": "LC2", "factor": 1.5}]}, ...]
    }
    """
    wb = Workbook()
    ws_first = wb.active
    wb.remove(ws_first)

    _sheet_general(wb, structural_model)
    _sheet_project(wb, structural_model)
    _sheet_materials(wb, structural_model)
    _sheet_cross_sections(wb, structural_model)
    _sheet_structural_nodes(wb, structural_model)
    _sheet_structural_members(wb, structural_model)
    _sheet_supports(wb, structural_model)
    _sheet_load_cases(wb, structural_model)
    _sheet_combinations(wb, structural_model)
    _sheet_loads(wb, structural_model)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_general(wb, model):
    ws = wb.create_sheet("General")
    rows = [
        ["SAF Version", SAF_VERSION],
        ["Generator", "BET Agent V2"],
        ["Generated At", datetime.utcnow().isoformat()],
        ["Units Length", "m"],
        ["Units Force", "kN"],
        ["Units Mass", "t"],
        ["Units Temperature", "C"],
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40


def _sheet_project(wb, model):
    ws = wb.create_sheet("Project")
    project = model.get("project", {})
    rows = [
        ["Name", project.get("name", "Projet BET")],
        ["Referentiel", project.get("referentiel", "sia")],
        ["Country", project.get("country", "CH")],
        ["Standard", "SIA 260-267" if project.get("referentiel", "sia") == "sia" else "Eurocodes"],
        ["ExposureClass", project.get("exposure_class", "XC2")],
        ["ConsequenceClass", project.get("consequence_class", "CC2")],
        ["SeismicZone", project.get("seismic_zone", "Z1b")],
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30


def _sheet_materials(wb, model):
    ws = wb.create_sheet("Materials")
    headers = ["Name", "Class", "Grade", "Density_kg_m3", "E_GPa", "fck_MPa", "fy_MPa", "fmk_MPa"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)

    # Collecte des matériaux utilisés dans le modèle
    used = set(m.get("material") for m in model.get("members") or [] if m.get("material"))
    used |= set(s.get("material") for s in model.get("surfaces") or [] if s.get("material"))
    # Toujours inclure les matériaux standards référencés
    if not used:
        used = {"C30/37", "S355"}

    for name in sorted(used):
        mat = MATERIALS_SIA.get(name, {"class": "Unknown", "grade": name})
        ws.append([
            name,
            mat.get("class"),
            mat.get("grade"),
            mat.get("density_kg_m3"),
            mat.get("E_GPa"),
            mat.get("fck_MPa"),
            mat.get("fy_MPa"),
            mat.get("fmk_MPa"),
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 16


def _sheet_cross_sections(wb, model):
    ws = wb.create_sheet("CrossSections")
    headers = ["Name", "Shape", "h_mm", "b_mm", "thickness_mm", "Material", "Usage"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)

    used = set(m.get("section") for m in model.get("members") or [] if m.get("section"))
    if not used:
        used = {"POU_30x50", "POT_30x30"}

    for name in sorted(used):
        cs = CROSS_SECTIONS_DEFAULTS.get(name, {"shape": "Rectangular"})
        ws.append([
            name,
            cs.get("shape"),
            cs.get("h_mm"),
            cs.get("b_mm"),
            cs.get("thickness_mm"),
            cs.get("material", ""),
            cs.get("usage", ""),
        ])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 14


def _sheet_structural_nodes(wb, model):
    ws = wb.create_sheet("StructuralNodes")
    headers = ["Name", "X_m", "Y_m", "Z_m"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    for node in model.get("nodes") or []:
        ws.append([
            node.get("id", ""),
            float(node.get("x", 0)),
            float(node.get("y", 0)),
            float(node.get("z", 0)),
        ])


def _sheet_structural_members(wb, model):
    ws = wb.create_sheet("Structural1DMembers")
    headers = ["Name", "Type", "NodeStart", "NodeEnd", "CrossSection", "Material", "Length_m"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    for m in model.get("members") or []:
        # Calcul longueur si nœuds présents
        length = m.get("length_m")
        if length is None:
            nodes = {n.get("id"): n for n in model.get("nodes") or []}
            ns = nodes.get(m.get("node_start"))
            ne = nodes.get(m.get("node_end"))
            if ns and ne:
                dx = ne.get("x", 0) - ns.get("x", 0)
                dy = ne.get("y", 0) - ns.get("y", 0)
                dz = ne.get("z", 0) - ns.get("z", 0)
                length = (dx ** 2 + dy ** 2 + dz ** 2) ** 0.5
        ws.append([
            m.get("id", ""),
            m.get("type", "beam"),
            m.get("node_start", ""),
            m.get("node_end", ""),
            m.get("section", ""),
            m.get("material", ""),
            round(length, 3) if length else None,
        ])


def _sheet_supports(wb, model):
    ws = wb.create_sheet("Supports")
    headers = ["Name", "Node", "Type", "RestraintUX", "RestraintUY", "RestraintUZ", "RestraintRX", "RestraintRY", "RestraintRZ"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    for s in model.get("supports") or []:
        t = s.get("type", "fixed")
        if t == "fixed":
            r = [True, True, True, True, True, True]
        elif t == "pinned":
            r = [True, True, True, False, False, False]
        elif t == "roller":
            r = [False, False, True, False, False, False]
        else:
            r = [True, True, True, False, False, False]
        ws.append([
            s.get("id", ""),
            s.get("node", ""),
            t,
            *r,
        ])


def _sheet_load_cases(wb, model):
    ws = wb.create_sheet("LoadCases")
    headers = ["Name", "Description", "ActionType", "LoadGroup"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    cases = model.get("load_cases") or []
    if not cases:
        # Défauts SIA 261
        cases = [
            {"id": "G1", "name": "Poids propre", "category": "Permanent"},
            {"id": "G2", "name": "Charges permanentes revêtements", "category": "Permanent"},
            {"id": "Q1", "name": "Charge d'exploitation (catégorie A)", "category": "Variable"},
            {"id": "QS", "name": "Neige", "category": "Variable"},
            {"id": "QW", "name": "Vent", "category": "Variable"},
        ]
    for c in cases:
        ws.append([
            c.get("id"),
            c.get("name"),
            c.get("category", "Variable"),
            c.get("group", "Main"),
        ])


def _sheet_combinations(wb, model):
    ws = wb.create_sheet("LoadCombinations")
    headers = ["Name", "Description", "Type", "Factors_json"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    combos = model.get("combinations") or []
    if not combos:
        # Combinaisons SIA 260 par défaut (indicatives)
        combos = [
            {"id": "ULS_STR", "name": "ELU STR : 1.35G + 1.5Q", "type": "linear_add",
             "factors": [{"case": "G1", "factor": 1.35}, {"case": "G2", "factor": 1.35},
                         {"case": "Q1", "factor": 1.5}]},
            {"id": "SLS_CHAR", "name": "ELS caractéristique : G + Q", "type": "linear_add",
             "factors": [{"case": "G1", "factor": 1.0}, {"case": "G2", "factor": 1.0},
                         {"case": "Q1", "factor": 1.0}]},
            {"id": "SLS_QP", "name": "ELS quasi-permanent : G + ψ2·Q", "type": "linear_add",
             "factors": [{"case": "G1", "factor": 1.0}, {"case": "G2", "factor": 1.0},
                         {"case": "Q1", "factor": 0.3}]},
        ]
    import json as _json
    for c in combos:
        ws.append([
            c.get("id"),
            c.get("name"),
            c.get("type", "linear_add"),
            _json.dumps(c.get("factors", []), ensure_ascii=False),
        ])
    ws.column_dimensions["D"].width = 60


def _sheet_loads(wb, model):
    ws = wb.create_sheet("Loads")
    headers = ["Name", "LoadCase", "Target", "TargetType", "Type", "Direction", "Value_kN_m_or_kN"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    for l in model.get("loads") or []:
        ws.append([
            l.get("id"),
            l.get("case"),
            l.get("target"),
            l.get("target_type", "member"),
            l.get("type", "uniform_vertical"),
            l.get("direction", "Z"),
            l.get("value_kN_m", l.get("value_kN", 0)),
        ])
