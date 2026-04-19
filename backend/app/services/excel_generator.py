"""Génération de fichiers Excel DPGF/DQE via openpyxl.

Formules actives, mise en forme professionnelle, en-tête, total calculé.
"""
import logging
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _apply_header_style(cell, fill_color: str = "1F2937") -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_cell_border(cell) -> None:
    thin = Side(border_style="thin", color="D4D4D4")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_dpgf_excel(
    project_name: str,
    lot: str,
    lines: list[dict[str, Any]],
    organization_name: str = "",
) -> bytes:
    """Génère un DPGF (Décomposition du Prix Global et Forfaitaire).

    lines: liste de {article, designation, unite, quantite, prix_unitaire, [sous_total]}.
    Calcul automatique des totaux par section et total général.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"DPGF {lot[:28]}"

    # En-tête
    ws["A1"] = f"DPGF — {lot.upper()}"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Projet : {project_name}"
    ws["A2"].font = Font(size=11)
    ws.merge_cells("A2:F2")

    ws["A3"] = f"Émis le {datetime.now().strftime('%d/%m/%Y')}{' — ' + organization_name if organization_name else ''}"
    ws["A3"].font = Font(size=9, color="737373")
    ws.merge_cells("A3:F3")

    # Ligne d'en-tête colonnes (ligne 5)
    headers = ["N° Article", "Désignation", "Unité", "Quantité", "P.U. HT (€)", "Total HT (€)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        _apply_header_style(cell)

    ws.row_dimensions[5].height = 28

    # Lignes
    row_num = 6
    first_data_row = row_num
    for line in lines:
        is_section = line.get("is_section", False)

        ws.cell(row=row_num, column=1, value=line.get("article", ""))
        ws.cell(row=row_num, column=2, value=line.get("designation", ""))

        if is_section:
            for c in range(1, 7):
                cell = ws.cell(row=row_num, column=c)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="525252", end_color="525252", fill_type="solid")
                _apply_cell_border(cell)
        else:
            ws.cell(row=row_num, column=3, value=line.get("unite", ""))
            ws.cell(row=row_num, column=4, value=line.get("quantite") or 0)
            ws.cell(row=row_num, column=5, value=line.get("prix_unitaire") or 0)
            # Formule calculée pour le sous-total
            ws.cell(row=row_num, column=6, value=f"=D{row_num}*E{row_num}")

            # Format nombres
            ws.cell(row=row_num, column=4).number_format = "#,##0.00"
            ws.cell(row=row_num, column=5).number_format = "#,##0.00 €"
            ws.cell(row=row_num, column=6).number_format = "#,##0.00 €"

            for c in range(1, 7):
                cell = ws.cell(row=row_num, column=c)
                _apply_cell_border(cell)
                if c == 2:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        row_num += 1

    # Total général (avec formule SUM)
    last_data_row = row_num - 1
    row_num += 1

    cell = ws.cell(row=row_num, column=5, value="TOTAL HT")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")

    total_cell = ws.cell(row=row_num, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})")
    total_cell.font = Font(bold=True, size=12)
    total_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    total_cell.number_format = "#,##0.00 €"
    _apply_cell_border(total_cell)

    row_num += 1
    cell = ws.cell(row=row_num, column=5, value="TVA 20%")
    cell.alignment = Alignment(horizontal="right")
    tva_cell = ws.cell(row=row_num, column=6, value=f"=F{row_num - 1}*0.2")
    tva_cell.number_format = "#,##0.00 €"

    row_num += 1
    cell = ws.cell(row=row_num, column=5, value="TOTAL TTC")
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal="right")
    ttc_cell = ws.cell(row=row_num, column=6, value=f"=F{row_num - 2}*1.2")
    ttc_cell.font = Font(bold=True, size=13)
    ttc_cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    ttc_cell.font = Font(bold=True, size=13, color="FFFFFF")
    ttc_cell.number_format = "#,##0.00 €"

    # Largeurs colonnes
    widths = {1: 12, 2: 55, 3: 8, 4: 12, 5: 14, 6: 16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze panes (fige ligne header)
    ws.freeze_panes = "A6"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_dqe_excel(
    project_name: str,
    lots_data: dict[str, list[dict]],
    organization_name: str = "",
) -> bytes:
    """Génère un DQE multi-lots (une feuille récapitulative + une par lot)."""
    wb = Workbook()

    # Feuille récapitulative
    ws_recap = wb.active
    ws_recap.title = "Récapitulatif"

    ws_recap["A1"] = "DQE — Décomposition par Lots"
    ws_recap["A1"].font = Font(bold=True, size=16)
    ws_recap.merge_cells("A1:C1")

    ws_recap["A2"] = f"Projet : {project_name}"
    ws_recap.merge_cells("A2:C2")

    ws_recap["A3"] = f"Émis le {datetime.now().strftime('%d/%m/%Y')}{' — ' + organization_name if organization_name else ''}"
    ws_recap["A3"].font = Font(size=9, color="737373")
    ws_recap.merge_cells("A3:C3")

    for idx, header in enumerate(["Lot", "Total HT (€)", "% du total"], start=1):
        cell = ws_recap.cell(row=5, column=idx, value=header)
        _apply_header_style(cell)

    row = 6
    total_formula_refs = []

    # Une feuille par lot
    for lot_name, lines in lots_data.items():
        sheet_name = lot_name[:31]
        ws_lot = wb.create_sheet(sheet_name)
        _fill_lot_sheet(ws_lot, project_name, lot_name, lines, organization_name)

        # Reference au total du lot dans la feuille récap
        # On suppose que le total HT est en F{last_row} — on calcule par SUM sur la plage
        last_line = 5 + len(lines) + 1  # header à ligne 5
        total_ref = f"'{sheet_name}'!F{last_line + 1}"  # "TOTAL HT" est 2 lignes après la dernière data
        # Pour éviter les casse-pieds, on recalcule:
        ws_recap.cell(row=row, column=1, value=lot_name)
        ws_recap.cell(row=row, column=2, value=f"=SUM('{sheet_name}'!F6:F{last_line})")
        ws_recap.cell(row=row, column=2).number_format = "#,##0.00 €"
        total_formula_refs.append(f"B{row}")
        row += 1

    # Total général
    if total_formula_refs:
        last_total_row = row
        ws_recap.cell(row=row, column=1, value="TOTAL HT").font = Font(bold=True, size=12)
        sum_formula = f"=SUM({','.join(total_formula_refs)})"
        total_cell = ws_recap.cell(row=row, column=2, value=sum_formula)
        total_cell.font = Font(bold=True, size=12)
        total_cell.number_format = "#,##0.00 €"
        total_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        # Pourcentages
        for idx, ref in enumerate(total_formula_refs):
            pct_row = 6 + idx
            pct_cell = ws_recap.cell(row=pct_row, column=3, value=f"={ref}/B{last_total_row}")
            pct_cell.number_format = "0.00%"

    ws_recap.column_dimensions["A"].width = 35
    ws_recap.column_dimensions["B"].width = 18
    ws_recap.column_dimensions["C"].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _fill_lot_sheet(ws, project_name: str, lot_name: str, lines: list[dict], organization_name: str) -> None:
    """Remplit une feuille de lot (utilitaire pour DQE)."""
    ws["A1"] = f"DQE — {lot_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Projet : {project_name}"
    ws.merge_cells("A2:F2")

    headers = ["N°", "Désignation", "Unité", "Quantité", "P.U. HT (€)", "Total HT (€)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        _apply_header_style(cell)

    row_num = 6
    first_data_row = row_num
    for line in lines:
        ws.cell(row=row_num, column=1, value=line.get("article", ""))
        ws.cell(row=row_num, column=2, value=line.get("designation", ""))
        ws.cell(row=row_num, column=3, value=line.get("unite", ""))
        ws.cell(row=row_num, column=4, value=line.get("quantite") or 0)
        ws.cell(row=row_num, column=5, value=line.get("prix_unitaire") or 0)
        ws.cell(row=row_num, column=6, value=f"=D{row_num}*E{row_num}")
        ws.cell(row=row_num, column=4).number_format = "#,##0.00"
        ws.cell(row=row_num, column=5).number_format = "#,##0.00 €"
        ws.cell(row=row_num, column=6).number_format = "#,##0.00 €"
        for c in range(1, 7):
            _apply_cell_border(ws.cell(row=row_num, column=c))
        row_num += 1

    # Total
    row_num += 1
    ws.cell(row=row_num, column=5, value="TOTAL HT").font = Font(bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="right")
    total = ws.cell(row=row_num, column=6, value=f"=SUM(F{first_data_row}:F{row_num - 2})")
    total.font = Font(bold=True)
    total.number_format = "#,##0.00 €"

    widths = {1: 10, 2: 50, 3: 8, 4: 12, 5: 14, 6: 16}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A6"
